from bs4 import BeautifulSoup
import requests
import re
import os
import time
from urllib.request import urlopen
from openpyxl import Workbook, load_workbook
import pandas as pd
from pyecharts import options as opts
from pyecharts.charts import Bar

# *-------------反反爬--------*

headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54",
           "Cookie":"__permanent_id=20221209153410146419057396291773630; ddscreen=2; __visit_id=20221219192038249375196318299854863; __out_refer=; dest_area=country_id%3D9000%26province_id%3D111%26city_id%3D0%26district_id%3D0%26town_id%3D0; pos_9_end=1671449099643; pos_0_end=1671449099805; ad_ids=20858015%7C%232; pos_0_start=1671449105644; secret_key=1ccfcc93e86cbca51f467086791652c6; __rpm=s_112100.94003212839%2C94003212840.1.1671450393965%7Clogin_page...1671450485829; sessionID=pc_a0c3083bc835ce9b4a757abc8c07d907dca844a302ecebade3d1e4a49e7883f; USERNUM=r+S7x+flyLA1d1icUEvpeg==; login.dangdang.com=.ASPXAUTH=DqH8QtCgJ33HmBD3BP29g1l4mm3/oi2cZDujbiw33po5lpigcXIUQg==; dangdang.com=email=NTI0YzY2YTViNjhkYTk0N0BkZG1vYmlsZV91c2VyLmNvbQ==&nickname=&display_id=5529400927094&customerid=qZ91KRWGHIsf/UkC7/gfPQ==&viptype=EceQdt7yVWE=&show_name=136****4890; ddoy=email=524c66a5b68da947@ddmobile_user.com&nickname=&validatedflag=2&uname=&utype=1&.ALFG=off&.ALTM=1671450485539; LOGIN_TIME=1671450488066; __trace_id=20221219194808076412254231099539885"
           }
def getPage(url):
    html = requests.get(url,headers=headers)
    bs = BeautifulSoup(html.text, 'html.parser')
    return bs

def getdata(bs):
    list = bs.select('#content > div > div.article > ol > li')
    for movie in list:
        #读取每个电影信息
        name = movie.select('a > span.title')
        #片名
        name = name[0].get_text()
        core = movie.select(' div.info > div.bd > div > span.rating_num')
        #评分
        core = core[0].get_text()
        corepeopel = movie.select('div.info > div.bd > div > span:nth-child(4)')
        #评分人数
        corepeopel = corepeopel[0].get_text()
        moviedata = movie.select(' div.info > div.bd > p:nth-child(1)')
        moviedata = moviedata[0].get_text()
        daoyandatas = re.finditer(r'导演:(?P<daoyan>.*?)主演', moviedata)
        for daoyan in daoyandatas:
            #导演
            daoyan = daoyan.group('daoyan')
        ypdata = re.findall(r'\d\d\d\d.*', moviedata)
        ypdata = ypdata[0]
        ypdata2 = re.finditer(r'(?P<year>.*?)/(?P<placs>.*?)/(?P<tipe>.*)', ypdata)
        for yp in ypdata2:
            # 上映年份
            year = yp.group("year").strip()
            # 地区
            placs = yp.group("placs")
            if '大陆' in placs or '香港' in placs or '台湾' in placs:
                placs = '中国'
            # 电影类型
            tipe = yp.group("tipe")
        #导入excel
        exceldata = []
        exceldata.append(name)
        exceldata.append(core)
        exceldata.append(corepeopel)
        exceldata.append(daoyan)
        exceldata.append(year)
        exceldata.append(placs)
        exceldata.append(tipe)
        excelpath = 'TOP250.xlsx'
        saveexcel(excelpath, exceldata)

def saveexcel(excelpath,exceldata):
    if not os.path.exists(excelpath):
        #若当前文件夹没有，则创建新Excel并命名
        tableTitle = ['片名','评分','评价人数','导演','上映年份','国家/地区','电影类型']
        wb = Workbook()
        ws = wb.active
        ws.title = 'sheet1'
        ws.append(tableTitle)
        wb.save(excelpath)
        time.sleep(3)
    #导入每个这部电影的数据
    wb = load_workbook(excelpath)
    ws = wb.active
    ws.title = 'sheet1'
    ws.append(exceldata)
    wb.save(excelpath)



# *--------------主函数-----------------*
def main():
    i = 1
    mainurl = 'https://movie.douban.com/top250'
    bs = getPage(mainurl)
    print("第{}页数据读取********************************************".format(i))
    getdata(bs)
    i = i + 1
    urllist = bs.select('#content > div > div.article > div.paginator > a')
    for urlin in urllist:
        url = mainurl + str(urlin.get("href"))
        bs = getPage(url)
        print("第{}页数据读取********************************************".format(i))
        getdata(bs)
        i = i + 1

    #用panda导出excel数据
    data = pd.read_excel('TOP250.xlsx')
    #第一图
    year_amount = data['上映年份'].value_counts()
    year_amount.columns = ['上映年份', '数量']
    year_amount = year_amount.sort_index()
    c = (
        Bar()
            .add_xaxis(list(year_amount.index))
            .add_yaxis('上映数量', year_amount.values.tolist())
            .set_global_opts(
            title_opts=opts.TitleOpts(title='各年份上映电影数量'),
            yaxis_opts=opts.AxisOpts(name='上映数量'),
            xaxis_opts=opts.AxisOpts(name='上映年份'),
            datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_='inside')], )
            .render('各年份上映电影数量.html')
    )
    #第二图
    placs_amount = data['国家/地区'].value_counts()
    placs_amount.columns = ['国家/地区', '数量']
    placs_amount = placs_amount.sort_values(ascending=True)
    c = (
        Bar()
        .add_xaxis(list(placs_amount.index)[-10:])
        .add_yaxis('地区上映数量', placs_amount.values.tolist()[-10:])
        .reversal_axis()
        .set_global_opts(
            title_opts=opts.TitleOpts(title='地区上映电影数量'),
            yaxis_opts=opts.AxisOpts(name='国家/地区'),
            xaxis_opts=opts.AxisOpts(name='上映数量'),
        )
        .set_series_opts(label_opts=opts.LabelOpts(position="right"))
        .render('各地区上映电影数量前十.html')
    )











if __name__ == '__main__': main()